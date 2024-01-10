import zipfile
import io
import pandas as pd
from openpyxl import load_workbook

def extract_and_convert_csv(zip_file, xlsx_file):
    try:
        # Extraer el archivo XLSX del ZIP en memoria
        with zipfile.ZipFile(zip_file, 'r') as zip_file:
            with zip_file.open(xlsx_file) as xlsx_file:
                excel_content = io.BytesIO(xlsx_file.read())
                workbook = load_workbook(excel_content)
                sheet = workbook.active

                # Deshacer la fusión de celdas antes de leer el archivo con pandas
                sheet.unmerge_cells('A4:A5') 
                sheet.unmerge_cells('B4:B5')

                # Copiar el valor de fusión en las nuevas celdas
                fusion_value = sheet['A4'].value
                sheet['A5'] = fusion_value
                sheet['B5'] = sheet['B4'].value

        # Leer el contenido del archivo Excel con pandas
        df = pd.read_excel(excel_content, skiprows=[0, 1, 2, 3, 5, 6, 12], usecols=lambda x: x not in ['Unnamed: 1'])

        # Eliminar el pie de página (últimas filas del DataFrame)
        footer_index = 1 
        df = df.iloc[:-footer_index]

        # Guardar el DataFrame como un archivo CSV
        csv_file = 'Victims_Age_by_Offense_Category_2022.csv'
        df.to_csv(csv_file, index=False)

        print(f'Archivo CSV creado con éxito: {csv_file}')
    except Exception as e:
        print(f'Error: {e}')