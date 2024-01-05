import zipfile
import io
import pandas as pd
from openpyxl import load_workbook, comments

# ZIP file path
zip_file = 'victims.zip'

# Name of the XLSX file inside the ZIP
xlsx_file = 'Victims_Age_by_Offense_Category_2022.xlsx'

try:
    # Extract the XLSX file from the ZIP in memory
    with zipfile.ZipFile(zip_file, 'r') as zip_file:
        with zip_file.open(xlsx_file) as xlsx_file:
            excel_content = io.BytesIO(xlsx_file.read())
            
            # Load the workbook with openpyxl
            workbook = load_workbook(excel_content)
            
            # Select the worksheet
            sheet = workbook.active
            
            # Merge cells before reading file with pandas
            sheet.unmerge_cells('A4:A5') 
            sheet.unmerge_cells('B4:B5')
            
            fusion_value = sheet['A4'].value
            sheet['A5'] = fusion_value
            sheet['B5'] = sheet['B4'].value


    df = pd.read_excel(excel_content, skiprows=[0, 1, 2, 3, 5, 6, 12], usecols=lambda x: x not in ['Unnamed: 1'])

    # Delete the footer (last rows of the DataFrame)
    footer_index = 1 
    df.iloc[3, 0]
    df = df.iloc[:-footer_index]

    # Save the DataFrame as a CSV
    csv_file = 'Victims_Age_by_Offense_Category_2022.csv'
    df.to_csv(csv_file, index=False)

    print(f'Archivo CSV creado con éxito: {csv_file}')
except Exception as e:
    print(f'Error: {e}')